import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import os

path = './result'

try:
    os.mkdir(path)
except FileExistsError:
    pass

search_text = input("Enter the search content : ")
page = int(input("How many page : "))

number = -9

title_list = []
link_list = []

print('\n')
print(f"{search_text}의 검색결과를 찾는 중입니다")
print(f"총 {page}페이지 검색")

for i in range(page):

    print(f"검색 페이지 수 : {i+1} / {page}")

    number += 10
    url = f'https://search.naver.com/search.naver?where=news&sm=tab_pge&query={search_text}&start={number}'
    content = requests.get(url)

    assert content.status_code == 200

    soup = BeautifulSoup(content.text, "html.parser")

    minseo = soup.find("ul", {"class": "list_news"})
    minseo = minseo.find_all("li")

    for i in minseo:
        try:
            title = i.find("a", {"class":"news_tit"}).text
            link = i.find("a", {"class":"news_tit"})
            title_list.append(title)
            link_list.append(link.get('href'))
        except:
            pass

print("검색 완료!")
print("파일에 저장중 . . .")

wb = Workbook()
ws = wb.active

row = 2

ws['A1'] = '뉴스명'
ws['B1'] = '링크'

for i, j in zip(title_list, link_list):
    ws['A' + str(row)] = i
    ws['B' + str(row)] = j
    row += 1

wb.save(f'{path}/{search_text}.xlsx')

print("파일에 저장 완료!")