import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

def file_write(content):
    f = open("test.txt",'w',encoding='UTF-8')

    if type(content) == list:
        for i in content:
            f.write(i)
    else:
        f.write(content)
        
    f.close()

# wb = Workbook()
# ws = wb.active

search_text = input("Search in Naver News: ")

url = f'https://search.naver.com/search.naver?where=news&sm=tab_pge&query={search_text}&sort=0&photo=0&field=0&pd=0&ds=&de=&cluster_rank=94&mynews=0&office_type=0&office_section_code=0&news_office_checked=&nso=so:r,p:all,a:all&start=1'
response = requests.get(url)

assert response.status_code == 200

newsParser = BeautifulSoup(response.content, "html.parser")
title = newsParser.select('.news_tit')

print(title[0])
# file_write(title)

# row = 2 * 10 - 10
# for lst in data:
#     title = lst.select('a.tit')[0].text
#     date = lst.select('span.time')[0].text
#     row = row + 1
#     ws['A' + str(row)] = title
#     ws['B' + str(row)] = date

# wb.save('test.xlsx')