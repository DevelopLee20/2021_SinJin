from time import time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import os
import datetime

# 변수 리스트
path = './result'
months = [31,28,31,30,31,30,31,31,30,31,30,31]
title_list = []
link_list = []
author_list = []
time_list = []
search_text = input("Enter the search content : ")
count = int(input("How many article : "))
page_num = 1
page = 1
skip = 0

# 기본 크롤링 루프
while True:
    url = f'https://search.naver.com/search.naver?where=news&sm=tab_pge&query={search_text}&start={page_num}'

    content = requests.get(url)
    assert content.status_code == 200

    contents = BeautifulSoup(content.text, "html.parser")
    news_list = contents.find("div",{"class":"group_news"})
    boxes = news_list.find_all("li",{"class":"bx"})

    # 검색 시간 저장
    now = datetime.datetime.now()
    n_year = int(now.year); n_month = int(now.month); n_day = int(now.day); n_hour = int(now.hour)

    for i in boxes:
        year = n_year; month = n_month; day = n_day; hour = n_hour

        tit = i.find("a",{"class":"news_tit"})
        title = tit.text
        author = i.find("a",{"class":"info"}).text

        if title in title_list and author_list[title_list.index(title)] == author:
            skip += 1
            continue

        link = tit.get('href')
        tm = i.find("span",{"class":"info"})
        tm = str(tm.text)
        
        if (tm.find("년")+1):
            year -= int(tm.split("년")[0])
        elif (tm.find("달")+1):
            month -= int(tm.split("달")[0])
        elif (tm.find("일")+1):
            day -= int(tm.split("일")[0])
        elif (tm.find("시")+1):
            hour -= int(tm.split("시")[0])
        else:
            pass

        if hour < 0:
            hour = 24 + hour
            day -= 1
        if day <= 0:
            month -= 1
            day = months[month-1] + day
        if month <= 0:
            year -= 1
            month = 12 + month

        time_str = str(year)+"-"+str(month)+"-"+str(day)

        if len(title_list) >= count:
            break

        author_list.append(author)
        time_list.append(time_str)
        title_list.append(title)
        link_list.append(link)

    
    print(f"| 페이지: {int(page_num/10)+1} | 찾은 수: {len(title_list)}/{count} | 중복 수: {skip} |")

    if len(title_list) >= count:
        break

    page_num += 10

# 저장 위치
try:
    os.mkdir(path)
except FileExistsError:
    pass

print("검색 완료!")
print("파일에 저장중 . . .")

# 파일 저장
wb = Workbook()
ws = wb.active

row = 2

ws['A1'] = '언론사'
ws['B1'] = '시간'
ws['C1'] = '제목'
ws['D1'] = '링크'

for i, j, p, q in zip(author_list, time_list, title_list, link_list):
    ws['A' + str(row)] = i
    ws['B' + str(row)] = j
    ws['C' + str(row)] = p
    ws['D' + str(row)] = q
    row += 1

wb.save(f'{path}/{search_text}.xlsx')

print("파일에 저장 완료!")