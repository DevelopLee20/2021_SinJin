# coding=utf-8
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

for page in range(1):
    result = page
    print "----------------------------------------"
    print "Trying to request " + str(result)

# 홈페이지에 요청하는 문구(주소만 수정가능)
response = requests.get('http://news.naver.com/main/search/search.nhn?query=%BD%C2%BA%CE%C1%B6%C0%DB&ie=MS949&x=0&y=0')

# 에러뜨면 알려주는 명령
assert response.status_code is 200
newrsParser = BeautifulSoup(response.content, "html.parser")
data = newrsParser.select('ul.srch_lst')
row = page * 10 - 10
for data in data:
    title = data.select('a.tit')[0].text
    date = data.select('span.time')[0].text
    row = row + 1
    ws['A' + str(row)] = title
    ws['B' + str(row)] = date


wb.save('test.xlsx')
